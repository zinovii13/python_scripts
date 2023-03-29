# %%
# import libs
import logging
from configparser import ConfigParser
import glob
import re
import os
from openpyxl import load_workbook
import pandas as pd
from pandas.errors import ParserError


# %%
# Read config.ini file
config_object = ConfigParser()
config_object.read(r"config/config.ini")


# %%
# setup log
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(
            filename=r"logs/log.log", mode="w"),
        logging.StreamHandler()]
)
logging.info(": Process start")


# setup csv delimiter
csv_delimiter = config_object["USER_SETTINGS"]["csv_delimiter"].strip()
if csv_delimiter == "":
    csv_delimiter = ";"
logging.info(": CSV_delimiter is %s", csv_delimiter)


# setup number of header rows
config_header_rows = config_object["USER_SETTINGS"]["header_rows"].strip()
try:
    header_rows = [
        row-1 for row in list(map(int, config_header_rows.split(",")))]
except:
    header_rows = [0]
logging.info(": Header rows - %s", config_header_rows)


# setup encoding type for csv file
encoding = config_object["USER_SETTINGS"]["encoding"].strip()
if encoding == "":
    encoding = "utf-8"
logging.info(": Encoding is %s", encoding)

# %%
# clean temp dir
for _ in glob.glob('temp/*'):
    os.remove(_)



# %%
# function change format excel_files to string
def change_format_to_text_in_excel(excel_filename):
    '''add docstring
    '''
    workbook = load_workbook(excel_filename)

    # format to text
    # for worksheet_name in workbook.sheetnames:
    for row in workbook.active.iter_rows():
        for cell in row:
            if cell.column_letter == "A" and cell.row > 1:
                workbook.active[cell.coordinate].number_format = "@"

    workbook.save(r'temp/' + os.path.basename(excel_filename))


# %%
# concat files into one
def function_concat_files_into_one():
    '''doctstring must add
    '''

    df = pd.DataFrame()

    for input_filename in glob.glob(r"01_in/**/*.*", recursive=True):
        if input_filename[-4:] == ".csv":
            try:
                input_df = pd.read_csv(input_filename, sep=csv_delimiter,
                                       header=header_rows, encoding=encoding, keep_default_na=False, dtype="string")
            except ParserError:
                logging.error(
                    ": Skipping incompatible file - %s", os.path.basename(input_filename))
        elif input_filename[-5:] == ".xlsx":
            change_format_to_text_in_excel(input_filename)
            input_df = pd.read_excel(
                r'temp/' + os.path.basename(input_filename), header=header_rows, keep_default_na=False, dtype="string")
        else:
            logging.error(
                ": Incorrect data type - %s", os.path.basename(input_filename))
            continue

        input_df.rename(columns=lambda x: re.sub(
            "Unnamed: [0-9]*_level_[0-9]", "", str(x)), inplace=True)
        input_df.rename(columns=lambda x: str(x).strip(), inplace=True)

        duplicate_cols = input_df.columns[input_df.columns.duplicated()]
        if len(duplicate_cols) > 0:
            logging.info(
                ": Duplicate column(s) - %s found in file %s" % (list(duplicate_cols), os.path.basename(input_filename)))
            input_df.columns = pd.io.parsers.base_parser.ParserBase({"names": list(
                input_df.columns), "usecols": None})._maybe_dedup_names(list(input_df.columns))
        input_df.insert(0, "FILENAME", f"{input_filename[6:]}")
        df = pd.concat([input_df, df])
    logging.info(": Concatenate file contains %s rows and %s columns",
                 df.shape[0], df.shape[1])
    if glob.glob(r"01_in/*")[0][-4:] == ".csv":
        df.to_csv(
            r"02_out/concat_file.csv", sep=csv_delimiter, index=False)

    elif glob.glob(r"01_in/*")[0][-5:] == ".xlsx":
        writer = pd.ExcelWriter(r"02_out/concat_file.xlsx")
        headers = pd.DataFrame(df.columns.tolist()).T
        headers.to_excel(writer, header=False, index=False)
        df.columns = pd.Index(range(len(df.columns)))  
        df.to_excel( writer, header=False, index=False, startrow=len(headers))
        writer.close()
    logging.info(": Process end")


# %%
function_concat_files_into_one()
